"""Lecture du fichier RH (.xlsx ou .csv) -> liste de fiches a importer.

Le fichier du RH a typiquement les colonnes : Matricule (ou « Mle »), Nom,
Prenom, « Solde de conge » (nombre a virgule decimale FR, ex. « 21,47 »).
On detecte les colonnes par leur en-tete (insensible a la casse/aux accents),
sur n'importe quelle des premieres lignes (l'en-tete n'est pas toujours en
ligne 1). Tres tolerant : une cellule illisible -> ignoree, jamais d'exception.
"""
import csv
import io
import unicodedata

from openpyxl import load_workbook


def _norm(s) -> str:
    """Minuscule, sans accents, espaces normalises (pour comparer des en-tetes)."""
    s = "" if s is None else str(s)
    s = "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )
    return " ".join(s.lower().split())


def _to_float(v) -> float:
    """« 21,47 » / « 21.47 » / 21.47 -> 21.47 ; vide/illisible -> 0.0."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "").replace(" ", "").replace(",", ".")
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


# Mots-cles d'en-tete -> champ. Ordre : on teste le plus specifique d'abord.
_FIELDS = {
    "matricule": ("matricule", "mle", "mat", "immatricule"),
    "prenom": ("prenom", "prenoms"),
    "nom": ("nom",),
    "solde_initial": ("solde", "solde de conge", "solde conge", "reliquat"),
    "poste": ("poste", "fonction"),
    "service": ("service", "departement", "departement / service"),
}


def _match_field(header: str):
    h = _norm(header)
    if not h:
        return None
    for field, keys in _FIELDS.items():
        if any(h == k or h.startswith(k) for k in keys):
            return field
    return None


def _map_header(cells) -> dict:
    """Ligne de cellules -> {champ: index}. Le 1er match gagne par champ."""
    mapping = {}
    for i, c in enumerate(cells):
        field = _match_field(c)
        if field and field not in mapping:
            mapping[field] = i
    return mapping


def _rows_from_xlsx(content: bytes):
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        yield list(row)
    wb.close()


def _rows_from_csv(content: bytes):
    text = content.decode("utf-8-sig", errors="replace")
    # Detecte le separateur (; courant en FR, sinon ,).
    sample = text[:2048]
    delim = ";" if sample.count(";") >= sample.count(",") else ","
    for row in csv.reader(io.StringIO(text), delimiter=delim):
        yield row


def parse(filename: str, content: bytes) -> list[dict]:
    """Renvoie [{matricule, nom, prenom, solde_initial, poste, service}, ...].

    Ignore les lignes sans matricule. `filename` sert juste a choisir le lecteur.
    """
    name = (filename or "").lower()
    rows = (_rows_from_csv(content) if name.endswith(".csv")
            else _rows_from_xlsx(content))

    mapping = None
    out = []
    for cells in rows:
        if mapping is None:
            m = _map_header(cells)
            # En-tete valable : on sait au moins ou est le matricule.
            if "matricule" in m and ("nom" in m or "solde_initial" in m):
                mapping = m
            continue
        if not cells:
            continue

        def _get(field):
            i = mapping.get(field)
            return cells[i] if i is not None and i < len(cells) else None

        matricule = _get("matricule")
        matricule = "" if matricule is None else str(matricule).strip()
        if not matricule:
            continue
        out.append({
            "matricule": matricule,
            "nom": (str(_get("nom")).strip() if _get("nom") is not None else None) or None,
            "prenom": (str(_get("prenom")).strip() if _get("prenom") is not None else None) or None,
            "solde_initial": _to_float(_get("solde_initial")),
            "poste": (str(_get("poste")).strip() if _get("poste") is not None else None) or None,
            "service": (str(_get("service")).strip() if _get("service") is not None else None) or None,
        })
    return out
