"""Export Excel (.xlsx) du calendrier — grille hebdo type Clockify, par mois.

Par agent : lignes = sa plage horaire perso (debut d'activite -> +9h, en heure
locale), colonnes = jours ouvres (Lun-Ven) regroupes par semaine sur tout le
mois. Chaque case = tous les projets captures cette heure-la (listes), couleur =
version dominante (V1/V2/V3/Autres), rouge si depassement du temps prevu.
On compte la PRESENCE sur le projet : actif + inactif (idle) ; la pause
volontaire est exclue (case laissee vide). Un commentaire par bloc-projet
(survol dans Excel/Sheets) donne Debut / Fin / Total.
Les statuts manuels (conge, attente, absence...) ne sont pas geres (non captures)
-> cases laissees vides. Heures en local Madagascar (UTC+3).
"""
from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import case, func, select

from app.common.enums import StateEnum
from app.database.database_session import SessionLocal
from app.database.models import Client, Employee, Project, Segment

_ACTIVE = StateEnum.ACTIVE.value
_IDLE = StateEnum.IDLE.value
_LOCAL_OFFSET = timedelta(hours=3)   # Madagascar UTC+3
_HOURS_SPAN = 10                     # nb de lignes d'heures par agent (debut -> +9)
_DEFAULT_START = 9                   # plage par defaut si aucune activite

# Teintes proches du PDF (hex sans '#').
_FILL = {
    "V1": "8E7CC3",      # violet
    "V2": "6FA8DC",      # bleu
    "V3": "EAD1DC",      # rose clair
    "AUTRES": "ECECEC",  # gris tres clair (quasi blanc)
    "OVER": "FF0000",    # rouge vif (depassement)
}
_DAYS = ["Lun", "Mar", "Mer", "Jeu", "Ven"]
_MONTHS = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet",
           "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
_NAVY = "0E2638"


def _gather(date_from, date_to, external_ids):
    """date_from/date_to : objets date (bornes locales incluses). external_ids :
    liste des collaborateurs (external_id) a inclure, ou None/[] = tous."""
    lo = datetime(date_from.year, date_from.month, date_from.day) - _LOCAL_OFFSET
    hi = (datetime(date_to.year, date_to.month, date_to.day, 23, 59, 59)
          - _LOCAL_OFFSET)
    with SessionLocal() as s:
        emp_q = select(Employee.id, Employee.external_id, Employee.name).where(
            Employee.is_active.is_(True))
        if external_ids:
            emp_q = emp_q.where(Employee.external_id.in_(external_ids))
        emp_rows = s.execute(emp_q.order_by(Employee.name)).all()
        employees = [{"id": r.id, "name": r.name or r.external_id} for r in emp_rows]
        emp_ids = [r.id for r in emp_rows]
        seg_q = (
            select(
                Segment.employee_id, Segment.project_id,
                Segment.started_at, Segment.ended_at, Segment.duration_sec,
                Project.video_name, Project.version,
                Project.estimated_duration_sec,
            )
            .join(Project, Segment.project_id == Project.id)
            .where(Segment.state.in_((_ACTIVE, _IDLE)),
                   Segment.started_at >= lo, Segment.started_at <= hi)
        )
        if emp_ids:
            seg_q = seg_q.where(Segment.employee_id.in_(emp_ids))
        segs = s.execute(seg_q).all()
        # Depassement : temps actif total (toutes periodes) > temps prevu.
        spent = {}
        for r in s.execute(
            select(Segment.project_id, Segment.duration_sec).where(
                Segment.state == _ACTIVE, Segment.project_id.is_not(None))
        ).all():
            spent[r.project_id] = spent.get(r.project_id, 0) + (r.duration_sec or 0)
    est = {r.project_id: (r.estimated_duration_sec or 0) for r in segs}
    overbudget = {pid for pid, e in est.items() if e > 0 and spent.get(pid, 0) > e}
    return employees, segs, overbudget


def _bucket(segs):
    """cells[(emp_id, date, hour)][label] = {sec, first, version, pid}."""
    cells = {}
    for r in segs:
        start = r.started_at + _LOCAL_OFFSET
        end = (r.ended_at + _LOCAL_OFFSET) if r.ended_at else (
            start + timedelta(seconds=r.duration_sec or 0))
        if end <= start:
            end = start + timedelta(seconds=max(r.duration_sec or 0, 1))
        cur = start
        while cur < end:
            seg_end = min(end, cur.replace(minute=0, second=0, microsecond=0)
                          + timedelta(hours=1))
            key = (r.employee_id, cur.date(), cur.hour)
            cell = cells.setdefault(key, {}).setdefault(
                r.video_name,
                {"sec": 0.0, "first": cur, "version": r.version, "pid": r.project_id},
            )
            cell["sec"] += (seg_end - cur).total_seconds()
            if cur < cell["first"]:
                cell["first"] = cur
            cur = seg_end
    return cells


def _fmt_hm(dt):
    return dt.strftime("%Hh%M")


def _fmt_dur(sec):
    h = int(sec // 3600)
    m = int((sec % 3600) // 60)
    return f"{h}h{m:02d}"


def _project_spans(segs):
    """{(emp_id, date_locale, label): {start, end, total}} sur actif + idle.

    Debut = 1er segment du jour, Fin = dernier, Total = somme des durees
    (la pause volontaire est deja exclue en amont dans _gather)."""
    spans = {}
    for r in segs:
        start = r.started_at + _LOCAL_OFFSET
        end = (r.ended_at + _LOCAL_OFFSET) if r.ended_at else (
            start + timedelta(seconds=r.duration_sec or 0))
        key = (r.employee_id, start.date(), r.video_name)
        agg = spans.get(key)
        if agg is None:
            spans[key] = {"start": start, "end": end,
                          "total": float(r.duration_sec or 0)}
        else:
            agg["total"] += float(r.duration_sec or 0)
            if start < agg["start"]:
                agg["start"] = start
            if end > agg["end"]:
                agg["end"] = end
    return spans


def _span_comment(label, info):
    return (f"{label}\n"
            f"Début : {_fmt_hm(info['start'])}\n"
            f"Fin : {_fmt_hm(info['end'])}\n"
            f"Total : {_fmt_dur(info['total'])}")


def _week_blocks(date_from, date_to):
    """[(num_semaine, [(date, dans_la_plage) x5 Lun..Ven]), ...] couvrant la
    plage [date_from, date_to]. Les jours hors plage sont marques (ombres)."""
    weeks = {}
    d = date_from
    while d <= date_to:
        if d.weekday() < 5:
            iso = d.isocalendar()
            weeks[(iso[0], iso[1])] = True
        d += timedelta(days=1)
    blocks = []
    for iy, iw in sorted(weeks):
        monday = date.fromisocalendar(iy, iw, 1)
        days = [(monday + timedelta(days=i),
                 date_from <= (monday + timedelta(days=i)) <= date_to)
                for i in range(5)]
        blocks.append((iw, days))
    return blocks


def build_calendar_xlsx(date_from, date_to, external_ids=None):
    employees, segs, overbudget = _gather(date_from, date_to, external_ids)
    cells = _bucket(segs)
    spans = _project_spans(segs)  # pour les commentaires Debut/Fin/Total
    blocks = _week_blocks(date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Calendrier"

    thin = Side(style="thin", color="D0D5DB")
    thick = Side(style="medium", color=_NAVY)  # separateur entre agents
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    head_fill = PatternFill("solid", fgColor=_NAVY)
    head_font = Font(color="FFFFFF", bold=True, size=10)
    agent_fill = PatternFill("solid", fgColor="F1F4F7")
    out_fill = PatternFill("solid", fgColor="F7F8FA")  # jour hors mois

    GAP = 1  # colonne vide etroite entre deux semaines
    n_weeks = len(blocks)

    def _wcol(k):  # colonne du 1er jour de la semaine k (0-indexee)
        return 3 + k * (5 + GAP)

    day_cols = [c for k in range(n_weeks) for c in range(_wcol(k), _wcol(k) + 5)]
    last_col = (_wcol(n_weeks - 1) + 4) if n_weeks else 2
    periode = f"du {date_from.strftime('%d/%m/%Y')} au {date_to.strftime('%d/%m/%Y')}"
    nb = f" — {len(employees)} collaborateur(s)"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    t = ws.cell(1, 1, f"Suivi du temps — {periode}{nb}")
    t.font = Font(bold=True, size=13, color=_NAVY)

    ROW_WEEK, ROW_DAY = 2, 3
    for col, label in ((1, "Agent"), (2, "Heure")):
        c = ws.cell(ROW_DAY, col, label)
        c.font = head_font; c.fill = head_fill; c.alignment = center; c.border = border
    for wi, (_iw, days) in enumerate(blocks, start=1):
        cstart = _wcol(wi - 1)
        ws.merge_cells(start_row=ROW_WEEK, start_column=cstart, end_row=ROW_WEEK,
                       end_column=cstart + 4)
        wc = ws.cell(ROW_WEEK, cstart, f"Semaine {wi}")
        wc.font = head_font; wc.fill = head_fill; wc.alignment = center
        for i, (dd, in_month) in enumerate(days):
            hc = ws.cell(ROW_DAY, cstart + i,
                         f"{_DAYS[i]} {dd.day:02d}" if in_month else "")
            hc.font = head_font; hc.fill = head_fill; hc.alignment = center
            hc.border = border

    row = ROW_DAY + 1
    last_block_end = ROW_DAY
    for emp in employees:
        hours = [h for (eid, _d, h) in cells if eid == emp["id"]]
        start = min(hours) if hours else _DEFAULT_START
        end = min(max(max(hours), start + _HOURS_SPAN - 1) if hours
                  else start + _HOURS_SPAN - 1, 23)
        hour_list = list(range(start, end + 1))
        block_start, block_end = row, row + len(hour_list) - 1

        ws.merge_cells(start_row=block_start, start_column=1,
                       end_row=block_end, end_column=1)
        ac = ws.cell(block_start, 1, emp["name"])
        ac.alignment = center; ac.fill = agent_fill
        ac.font = Font(bold=True, size=10, color=_NAVY)

        commented = set()  # (date, label) deja commentes pour cet agent
        for idx, hour in enumerate(hour_list):
            r = block_start + idx
            hcell = ws.cell(r, 2, hour)  # juste le nombre (centré, sans « h »)
            hcell.alignment = center; hcell.border = border
            hcell.font = Font(size=9, color="6C7884", bold=True)
            for k, (_iw, days) in enumerate(blocks):
                cstart = _wcol(k)
                for i, (dd, in_month) in enumerate(days):
                    cc = ws.cell(r, cstart + i)
                    cc.alignment = left; cc.border = border
                    if not in_month:
                        cc.fill = out_fill
                        continue
                    bucket = cells.get((emp["id"], dd, hour))
                    if not bucket:
                        continue
                    items = sorted(bucket.items(), key=lambda kv: kv[1]["first"])
                    cc.value = "\n".join(lbl for lbl, _ in items)
                    dom = max(bucket.values(), key=lambda v: v["sec"])
                    color = (_FILL["OVER"] if dom["pid"] in overbudget
                             else _FILL.get((dom["version"] or "").upper(),
                                            _FILL["AUTRES"]))
                    cc.fill = PatternFill("solid", fgColor=color)
                    # Commentaire Debut/Fin/Total, une seule fois par projet-jour
                    # (attache a sa 1ere case, en haut de son bloc).
                    lines = []
                    for lbl, _ in items:
                        ck = (dd, lbl)
                        if ck in commented:
                            continue
                        commented.add(ck)
                        info = spans.get((emp["id"], dd, lbl))
                        if info:
                            lines.append(_span_comment(lbl, info))
                    if lines:
                        note = Comment("\n\n".join(lines), "Calendrier")
                        note.width, note.height = 230, 120
                        cc.comment = note

        # Fusion verticale des heures consecutives identiques (par colonne jour).
        for c_idx in day_cols:
            rr = block_start
            while rr <= block_end:
                val = ws.cell(rr, c_idx).value
                if val:
                    rr2 = rr
                    while rr2 + 1 <= block_end and ws.cell(rr2 + 1, c_idx).value == val:
                        rr2 += 1
                    if rr2 > rr:
                        ws.merge_cells(start_row=rr, start_column=c_idx,
                                       end_row=rr2, end_column=c_idx)
                    rr = rr2 + 1
                else:
                    rr += 1

        # Ligne EPAISSE au-dessus du bloc -> separation nette entre agents.
        top_sep = Border(top=thick, bottom=thin, left=thin, right=thin)
        for c_idx in (1, 2, *day_cols):
            ws.cell(block_start, c_idx).border = top_sep
        last_block_end = block_end
        row = block_end + 1

    # Ligne epaisse de fermeture sous le dernier agent.
    if employees:
        for c_idx in (1, 2, *day_cols):
            cur = ws.cell(last_block_end, c_idx)
            keep_top = cur.border.top if cur.border else thin
            cur.border = Border(bottom=thick, top=keep_top, left=thin, right=thin)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 6
    day_set = set(day_cols)
    for ci in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = (
            22 if ci in day_set else 2.5)  # gap = colonne etroite

    # Legende (verticale) : libelle (col A) + pastille couleur (col B), par ligne.
    leg = row + 1
    ws.merge_cells(start_row=leg, start_column=1, end_row=leg, end_column=2)
    lt = ws.cell(leg, 1, "")  # case conservée, sans le texte « Légende »
    lt.alignment = center
    lt.border = border
    ws.cell(leg, 2).border = border
    for i, (lab, key) in enumerate(
        [("V1", "V1"), ("V2", "V2"), ("V3", "V3"),
         ("Autres", "AUTRES"), ("Dépassement timer", "OVER")]
    ):
        r = leg + 1 + i
        lc = ws.cell(r, 1, lab)
        lc.font = Font(size=9); lc.border = border; lc.alignment = center
        sw = ws.cell(r, 2)
        sw.fill = PatternFill("solid", fgColor=_FILL[key]); sw.border = border

    ws.freeze_panes = "C4"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _spent_by_project():
    """project_id -> temps actif cumulé (toutes périodes) = « Temps actuel » de la
    page Projets (même calcul que projects._spent_lookup)."""
    with SessionLocal() as s:
        rows = s.execute(
            select(
                Segment.project_id,
                func.sum(
                    case((Segment.state == _ACTIVE, Segment.duration_sec), else_=0)
                ),
            )
            .where(Segment.project_id.is_not(None))
            .group_by(Segment.project_id)
        ).all()
    return {r[0]: (r[1] or 0) for r in rows}


def build_recap_xlsx(date_from, date_to, external_ids=None):
    """Récap par collaborateur : ses projets travaillés sur la plage, avec
    Temps prévu / Temps actuel (cumul actif, comme la page Projets) / Temps
    restant (= prévu − actuel ; négatif = dépassé)."""
    lo = datetime(date_from.year, date_from.month, date_from.day) - _LOCAL_OFFSET
    hi = (datetime(date_to.year, date_to.month, date_to.day, 23, 59, 59)
          - _LOCAL_OFFSET)
    spent = _spent_by_project()
    with SessionLocal() as s:
        emp_q = select(Employee.id, Employee.external_id, Employee.name).where(
            Employee.is_active.is_(True))
        if external_ids:
            emp_q = emp_q.where(Employee.external_id.in_(external_ids))
        emp_rows = s.execute(emp_q.order_by(Employee.name)).all()
        emp_ids = [r.id for r in emp_rows]

        # (collaborateur, projet) où il a eu de l'activité ACTIVE sur la plage
        # (même critère que le calendrier, pour lister exactement pareil).
        # project_id NULL -> « (non identifié) » : on ne masque rien.
        pair_q = (
            select(
                Segment.employee_id, Segment.project_id,
                Project.video_name, Project.version,
                Project.estimated_duration_sec, Client.name.label("client"),
            )
            .join(Project, Segment.project_id == Project.id, isouter=True)
            .join(Client, Project.client_id == Client.id, isouter=True)
            .where(Segment.state == _ACTIVE,
                   Segment.started_at >= lo, Segment.started_at <= hi)
        )
        if emp_ids:
            pair_q = pair_q.where(Segment.employee_id.in_(emp_ids))
        worked = {}  # emp_id -> {project_id: {video, version, client, est}}
        for r in s.execute(pair_q).all():
            byp = worked.setdefault(r.employee_id, {})
            byp.setdefault(r.project_id, {
                "video": r.video_name or "(non identifié)",
                "version": r.version or "",
                "client": r.client or "",
                "est": r.estimated_duration_sec or 0,
            })

    wb = Workbook()
    ws = wb.active
    ws.title = "Récap"

    thin = Side(style="thin", color="D0D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    head_fill = PatternFill("solid", fgColor=_NAVY)
    head_font = Font(color="FFFFFF", bold=True, size=10)
    name_fill = PatternFill("solid", fgColor="F1F4F7")
    red = Font(size=10, bold=True, color="DC2626")
    green = Font(size=10, bold=True, color="059669")

    headers = ["Collaborateur", "Projet", "Version", "Client",
               "Temps prévu", "Temps restant"]
    last_col = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    periode = f"du {date_from.strftime('%d/%m/%Y')} au {date_to.strftime('%d/%m/%Y')}"
    t = ws.cell(1, 1, f"Récap prévu / actuel / restant — {periode}")
    t.font = Font(bold=True, size=13, color=_NAVY)

    for col, h in enumerate(headers, start=1):
        c = ws.cell(3, col, h)
        c.font = head_font; c.fill = head_fill
        c.alignment = center; c.border = border

    row = 4
    for er in emp_rows:
        projs = worked.get(er.id)
        if not projs:
            continue
        start = row
        for pid, info in sorted(
            projs.items(),
            key=lambda kv: (kv[1]["video"] == "(non identifié)",
                            (kv[1]["video"] or "").lower()),
        ):
            est = info["est"]
            sp = spent.get(pid, 0)
            ws.cell(row, 2, info["video"]).alignment = left
            ws.cell(row, 3, info["version"]).alignment = center
            ws.cell(row, 4, info["client"]).alignment = left
            ws.cell(row, 5, _fmt_dur(est) if est else "-").alignment = center
            rc = ws.cell(row, 6)
            rc.alignment = center
            if est:
                rem = est - sp
                rc.value = f"-{_fmt_dur(-rem)}" if rem < 0 else _fmt_dur(rem)
                rc.font = red if rem < 0 else green
            else:
                rc.value = "-"
            for col in range(2, last_col + 1):
                ws.cell(row, col).border = border
            row += 1
        # Nom du collaborateur fusionné sur ses lignes.
        ws.merge_cells(start_row=start, start_column=1, end_row=row - 1, end_column=1)
        nc = ws.cell(start, 1, er.name or er.external_id)
        nc.font = Font(bold=True, size=10, color=_NAVY)
        nc.fill = name_fill
        nc.alignment = center
        for rr in range(start, row):
            ws.cell(rr, 1).border = border

    widths = [18, 32, 7, 20, 12, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
